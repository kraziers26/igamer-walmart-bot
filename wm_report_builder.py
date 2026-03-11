"""
wm_report_builder.py — Walmart Excel Report Builder
iGamer Corp | Mirrors report_builder.py exactly.
Only differences:
  - "Best Buy" → "Walmart" in all titles/labels
  - "BB Rank"  → "WM Rank"
  - "BB"       → "WM" in filenames
  - priceUpdateDate falls back to drop_date from cache (set by wm_fetcher)
  - Deal Age label references cache-based detection, not native API field
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import pytz

# ── Palette ───────────────────────────────────────────────────────────────────
C_NAVY        = "1F4E79"
C_DARK_GREEN  = "1E5631"
C_MID_BLUE    = "2E75B6"
C_GREEN_BG    = "E2EFDA"
C_GREEN_DARK  = "375623"
C_YELLOW_BG   = "FFF2CC"
C_YELLOW_DARK = "7F6000"
C_RED_BG      = "FCE4D6"
C_RED_DARK    = "843C0C"
C_WHITE       = "FFFFFF"
C_GREY_HDR    = "D9E1F2"
C_ORANGE      = "ED7D31"
C_LINK        = "0563C1"
C_SUMMARY_ALT = "EBF3FB"
C_WM_BLUE     = "0071CE"   # Walmart brand blue — used as accent

THIN        = Side(style="thin",   color="CCCCCC")
MED         = Side(style="medium", color="AAAAAA")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr_font(size=10, bold=True, color=C_WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=9, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(color):
    return PatternFill("solid", fgColor=color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap, indent=1)


# ── Signal scoring ────────────────────────────────────────────────────────────

def signal_score(p: dict) -> int:
    """Fresh Deal Score — uses pre-computed fresh_score set by wm_fetcher."""
    if "fresh_score" in p:
        return p["fresh_score"]
    # Fallback inline calculation (mirrors wm_fetcher logic)
    score = 0
    price_date = p.get("priceUpdateDate")   # set to drop_date by wm_fetcher
    if price_date:
        try:
            dt   = datetime.fromisoformat(price_date)
            now  = datetime.now(dt.tzinfo) if dt.tzinfo else datetime.utcnow()
            days = (now - dt).days
            if days == 0:    score += 4
            elif days <= 2:  score += 3
            elif days <= 7:  score += 1
        except Exception:
            pass
    if p.get("onSale"):     score += 2
    pct = float(p.get("percentSavings") or 0)
    if pct >= 20:   score += 3
    elif pct >= 10: score += 2
    elif pct >= 5:  score += 1
    save_d = float(p.get("dollarSavings") or 0)
    if save_d >= 300:   score += 2
    elif save_d >= 100: score += 1
    bs = p.get("bestSellingRank")
    if bs and bs <= 500: score += 1
    return score

def hot_label(score: int) -> str:
    if score >= 9:  return "🔴 HOT BUY"
    if score >= 6:  return "🟠 Strong"
    if score >= 3:  return "🟡 Moderate"
    return "⚪ Watch"

def row_bg(p: dict) -> str:
    if not p.get("onlineAvailability", True): return "EEEEEE"
    sc = signal_score(p)
    if sc >= 9: return "FFE8E8"
    pct = float(p.get("percentSavings") or 0)
    if pct >= 15: return C_GREEN_BG
    if pct >= 5:  return C_YELLOW_BG
    return C_WHITE

def deal_age(price_update_date) -> str:
    """
    Human-readable deal age.
    For Walmart this is the drop_date from our price cache, not a native API field.
    Label reflects that — 'Detected today' instead of 'New today'.
    """
    if not price_update_date:
        return "— No history yet"
    try:
        dt   = datetime.fromisoformat(str(price_update_date))
        now  = datetime.now(dt.tzinfo) if dt.tzinfo else datetime.utcnow()
        days = (now - dt).days
        if days == 0:  return "🟢 Detected today"
        if days <= 2:  return f"🟢 {days}d ago"
        if days <= 7:  return f"🟡 {days}d active"
        if days <= 14: return f"🟠 {days}d aging"
        return               f"🔴 {days}d old"
    except Exception:
        return "—"


# ── Header helpers ────────────────────────────────────────────────────────────

def write_title_rows(ws, title: str, ts: str, ncols: int):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value     = f"iGamer Corp  •  Walmart Market Intelligence  •  {title}"
    c.font      = Font(name="Arial", size=14, bold=True, color=C_WHITE)
    c.fill      = fill(C_WM_BLUE)
    c.alignment = left()
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]
    c.value     = (
        f"Generated: {ts}   •   "
        "🟢 Fresh Deal = price drop detected by daily cache comparison   "
        "🛒 WM Rank = Walmart sales rank   •   "
        "🔴 HOT BUY = fresh drop + deep discount"
    )
    c.font      = Font(name="Arial", size=9, italic=True, color=C_WHITE)
    c.fill      = fill(C_NAVY)
    c.alignment = left()
    ws.row_dimensions[2].height = 16


def write_col_headers(ws, headers_widths: list, row: int, hdr_color: str = C_DARK_GREEN):
    for col, (h, w) in enumerate(headers_widths, 1):
        c           = ws.cell(row=row, column=col, value=h)
        c.font      = hdr_font(9)
        c.fill      = fill(hdr_color)
        c.alignment = center(wrap=True)
        c.border    = BORDER_THIN
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 30


# ── Category sheet ────────────────────────────────────────────────────────────

CAT_HEADERS = [
    ("RANK",            5),
    ("BRAND",           10),
    ("PRODUCT NAME",    55),
    ("SALE PRICE",      12),
    ("REG PRICE",       12),
    ("SAVE $",          10),
    ("SAVE %",          10),
    ("ON SALE?",        10),
    ("IN STOCK?",       10),
    ("🛒 WM RANK",      14),
    ("FRESH SCORE",     12),
    ("SIGNAL",          13),
    ("DEAL AGE",        18),
    ("BUY LINK",        16),
]

def build_category_sheet(wb, cat_name: str, products: list, ts: str):
    ws    = wb.create_sheet(cat_name)
    ncols = len(CAT_HEADERS)
    write_title_rows(ws, cat_name, ts, ncols)
    write_col_headers(ws, CAT_HEADERS, row=3)

    for rank, p in enumerate(products, 1):
        r  = 3 + rank
        bg = row_bg(p)
        sc = signal_score(p)

        sale_price = float(p.get("salePrice") or 0)
        reg_price  = float(p.get("regularPrice") or sale_price)
        save_d     = float(p.get("dollarSavings") or 0)
        save_pct   = float(p.get("percentSavings") or 0)
        on_sale    = bool(p.get("onSale"))
        in_stock   = bool(p.get("onlineAvailability", True))
        brand      = p.get("manufacturer", "—")
        name       = p.get("name", "—")
        url        = p.get("url", "")
        age        = deal_age(p.get("priceUpdateDate"))

        vals = [
            rank, brand, name,
            sale_price, reg_price if reg_price != sale_price else "—",
            f"${save_d:.2f}" if save_d > 0 else "—",
            f"{save_pct:.1f}%" if save_pct > 0 else "—",
            "✅ Yes" if on_sale else "❌ No",
            "✅ Yes" if in_stock else "❌ No",
            p.get("best_seller_str", "—"),
            f"{sc}/13",
            hot_label(sc),
            age,
            "🛒 Buy Now",
        ]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill   = fill(bg)
            c.border = BORDER_THIN

            if col == 14:  # buy link
                if url:
                    c.hyperlink = url
                c.font      = Font(name="Arial", size=9, bold=True,
                                   color=C_LINK, underline="single")
                c.alignment = center()
            elif col == 3:  # product name
                c.font      = body_font(9)
                c.alignment = left(wrap=True)
            elif col == 4:  # sale price
                c.number_format = "$#,##0.00"
                c.font = Font(name="Arial", size=9, bold=True,
                              color=C_GREEN_DARK if on_sale else "000000")
                c.alignment = center()
            elif col == 5:  # reg price
                c.number_format = "$#,##0.00" if isinstance(val, float) else "@"
                c.font      = body_font(9)
                c.alignment = center()
            elif col in (6, 7):  # save $ / %
                c.font = Font(name="Arial", size=9, bold=True,
                              color=C_GREEN_DARK if save_pct >= 15 else
                              (C_YELLOW_DARK if save_pct > 0 else "888888"))
                c.alignment = center()
            elif col == 12:  # signal label
                c.font = Font(name="Arial", size=9, bold=True,
                              color=C_RED_DARK if sc >= 9 else
                              (C_ORANGE if sc >= 6 else "555555"))
                c.alignment = center()
            elif col == 13:  # deal age
                c.font      = body_font(9)
                c.alignment = center()
            else:
                c.font      = body_font(9, bold=(col == 2))
                c.alignment = center() if col != 3 else left()

        ws.row_dimensions[r].height = 36

    # Footer
    foot_row = 3 + len(products) + 1
    ws.merge_cells(f"A{foot_row}:{get_column_letter(ncols)}{foot_row}")
    on_sale_count = sum(1 for p in products if p.get("onSale"))
    hot_count     = sum(1 for p in products if signal_score(p) >= 9)
    best_pct      = max((float(p.get("percentSavings") or 0) for p in products), default=0)
    c = ws.cell(row=foot_row, column=1,
                value=f"✅ {on_sale_count}/{len(products)} on sale   "
                      f"🔴 {hot_count} hot buys   "
                      f"💰 Best discount: {best_pct:.1f}%   "
                      f"📦 {len(products)} products shown")
    c.font      = hdr_font(9)
    c.fill      = fill(C_WM_BLUE)
    c.alignment = left()
    ws.row_dimensions[foot_row].height = 20

    ws.freeze_panes = ws.cell(row=4, column=1)
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{3 + len(products)}"


# ── Summary sheet ─────────────────────────────────────────────────────────────

SUMMARY_NCOLS = 12

def section_hdr(ws, row: int, text: str, color: str = C_NAVY) -> int:
    ws.merge_cells(f"A{row}:{get_column_letter(SUMMARY_NCOLS)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = hdr_font(10)
    c.fill      = fill(color)
    c.alignment = left()
    ws.row_dimensions[row].height = 22
    return row + 1


def build_summary_sheet(wb, all_data: dict, ts: str, filter_key: str = "full"):
    ws = wb.create_sheet("📊 SUMMARY", 0)

    widths = [24, 10, 10, 14, 14, 10, 14, 12, 12, 14, 18, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells(f"A1:{get_column_letter(SUMMARY_NCOLS)}1")
    c = ws["A1"]
    c.value     = "iGamer Corp  —  Walmart Daily Market Intelligence"
    c.font      = Font(name="Arial", size=16, bold=True, color=C_WHITE)
    c.fill      = fill(C_WM_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{get_column_letter(SUMMARY_NCOLS)}2")
    c = ws["A2"]
    filter_labels = {
        "full":     "📦 Established Deals",
        "trending": "🆕 Fresh Deals",
        "selling":  "🛒 Best Sellers",
        "on_sale":  "💰 On Sale Only",
        "hot":      "🔴 HOT BUYS Only",
    }
    filter_label = filter_labels.get(filter_key, "⚡ Full Report")
    c.value     = (
        f"Generated: {ts}   •   Filter: {filter_label}   •   "
        "Live data from Walmart.com  •  Freshness via daily price cache"
    )
    c.font      = Font(name="Arial", size=9, italic=True, color=C_WHITE)
    c.fill      = fill(C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    ROW = 4

    # ── Section 1: Category Overview ─────────────────────────────────────────
    ROW = section_hdr(ws, ROW, "  📋  CATEGORY OVERVIEW")

    ov_headers = [
        ("CATEGORY", 14), ("PRODUCTS", 10), ("ON SALE", 10),
        ("BEST DISC $", 13), ("BEST DISC %", 13), ("🔴 HOT BUYS", 12),
        ("AVG PRICE", 13), ("TOP BRAND", 12), ("GO TO SHEET", 14),
    ]
    for col, (h, _) in enumerate(ov_headers, 1):
        c           = ws.cell(row=ROW, column=col, value=h)
        c.font      = hdr_font(9)
        c.fill      = fill("2C3E6B")
        c.alignment = center(wrap=True)
        c.border    = BORDER_THIN
    ws.row_dimensions[ROW].height = 28
    ROW += 1

    for cat_name, cat_data in all_data.items():
        products  = cat_data["products"]
        on_sale   = sum(1 for p in products if p.get("onSale"))
        best_d    = max((float(p.get("dollarSavings") or 0) for p in products), default=0)
        best_pct  = max((float(p.get("percentSavings") or 0) for p in products), default=0)
        hot       = sum(1 for p in products if signal_score(p) >= 9)
        avg_price = sum(float(p.get("salePrice") or 0) for p in products) / max(len(products), 1)
        brands    = {}
        for p in products:
            b = p.get("manufacturer", "?")
            brands[b] = brands.get(b, 0) + 1
        top_brand = max(brands, key=brands.get) if brands else "—"

        alt = ROW % 2 == 0
        bg  = C_SUMMARY_ALT if alt else C_WHITE
        row_vals = [
            cat_name, len(products), f"{on_sale}/{len(products)}",
            f"${best_d:.0f}" if best_d else "—",
            f"{best_pct:.1f}%" if best_pct else "—",
            f"🔴 {hot}" if hot else "None",
            avg_price, top_brand, f"→ {cat_name}",
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=ROW, column=col, value=val)
            c.fill = fill(bg); c.border = BORDER_THIN
            if col == 7:
                c.number_format = "$#,##0.00"
                c.font = body_font(9, bold=True)
            elif col == 6 and hot:
                c.font = Font(name="Arial", size=9, bold=True, color=C_RED_DARK)
            elif col in (4, 5) and (best_d or best_pct):
                c.font = Font(name="Arial", size=9, bold=True, color=C_GREEN_DARK)
            elif col == 9:
                c.hyperlink = f"#{cat_name}!A1"
                c.font = Font(name="Arial", size=9, color=C_LINK, underline="single")
            else:
                c.font = body_font(9, bold=(col == 1))
            c.alignment = center() if col != 1 else left()
        ws.row_dimensions[ROW].height = 20
        ROW += 1

    ROW += 1

    # ── Section 2: Top Deals ──────────────────────────────────────────────────
    ROW = section_hdr(ws, ROW, "  🏆  TODAY'S TOP DEALS  —  Ranked by deal freshness + discount depth", "1E5631")

    deal_headers = [
        ("#", 4), ("TIER", 14), ("BRAND", 12), ("CATEGORY", 16), ("PRODUCT", 44),
        ("PRICE", 11), ("SAVE $", 9), ("SAVE %", 9),
        ("SCORE", 10), ("SIGNALS", 22), ("BUY", 12), ("WHY", 32),
    ]
    for col, (h, _) in enumerate(deal_headers, 1):
        c           = ws.cell(row=ROW, column=col, value=h)
        c.font      = hdr_font(9)
        c.fill      = fill("1E5631")
        c.alignment = center(wrap=True)
        c.border    = BORDER_THIN
    ws.row_dimensions[ROW].height = 28
    ROW += 1

    # Collect and rank all products
    all_products = []
    for cat_name, cat_data in all_data.items():
        for p in cat_data["products"]:
            pp = dict(p)
            pp["_cat"]   = cat_name
            pp["_score"] = signal_score(p)
            pp["_pct"]   = float(p.get("percentSavings") or 0)
            all_products.append(pp)

    all_products.sort(key=lambda x: (x["_score"], x["_pct"]), reverse=True)

    must_act   = [p for p in all_products if p["_score"] >= 9 and p.get("onSale")][:5]
    worth_look = [p for p in all_products if 5 <= p["_score"] < 9 and p.get("onSale")][:6]
    worth_look += [p for p in all_products
                   if p["_score"] >= 6 and not p.get("onSale")
                   and p not in worth_look][:3]
    worth_look = worth_look[:6]

    def signals_str(p):
        parts = []
        bs_str = p.get("best_seller_str", "")
        if bs_str and bs_str != "—": parts.append(bs_str)
        pct = float(p.get("percentSavings") or 0)
        if pct > 0: parts.append(f"💰 {pct:.0f}% off")
        if p.get("onSale"): parts.append("On Sale")
        fl = p.get("freshness_label") or ""
        if fl and fl != "—": parts.append(fl)
        return " • ".join(parts) or "—"

    def why_str(p):
        parts  = []
        sc     = p["_score"]
        pct    = p["_pct"]
        save_d = float(p.get("dollarSavings") or 0)
        bs_str = p.get("best_seller_str", "")
        fl     = p.get("freshness_label", "")
        if sc >= 9:   parts.append("Fresh drop + deep cut")
        elif sc >= 6: parts.append("Good deal, act soon")
        if pct >= 20:   parts.append(f"{pct:.0f}% off — deep cut")
        elif pct >= 10: parts.append(f"{pct:.0f}% off")
        if save_d >= 200: parts.append(f"${save_d:.0f} saved")
        if bs_str and bs_str != "—": parts.append(bs_str)
        if fl and ("today" in fl.lower() or "new" in fl.lower()): parts.append("Price just dropped")
        if not p.get("onlineAvailability", True): parts.append("⚠️ check stock")
        return " | ".join(parts) or "—"

    deal_num = 1
    for tier_lbl, tier_txt_c, tier_deals in [
        ("🔴 MUST ACT",   C_RED_DARK,    must_act),
        ("🟠 WORTH LOOK", C_YELLOW_DARK, worth_look),
    ]:
        for p in tier_deals:
            sc       = p["_score"]
            bg       = C_RED_BG if sc >= 9 else C_YELLOW_BG
            save_d   = float(p.get("dollarSavings") or 0)
            save_pct = p["_pct"]
            url      = p.get("url", "")

            row_vals = [
                deal_num, tier_lbl, p.get("manufacturer", "—"), p["_cat"], p.get("name", "—"),
                float(p.get("salePrice") or 0),
                f"${save_d:.2f}" if save_d else "—",
                f"{save_pct:.1f}%" if save_pct else "—",
                f"{sc}/13",
                signals_str(p), "🛒 Buy Now", why_str(p),
            ]
            for col, val in enumerate(row_vals, 1):
                c = ws.cell(row=ROW, column=col, value=val)
                c.fill = fill(bg); c.border = BORDER_THIN
                if col == 2:
                    c.font = Font(name="Arial", size=9, bold=True, color=tier_txt_c)
                    c.alignment = center()
                elif col == 3:
                    c.font = Font(name="Arial", size=9, bold=True, color=C_NAVY)
                    c.alignment = center()
                elif col == 5:
                    c.font = body_font(9, bold=True)
                    c.alignment = left(wrap=True)
                elif col == 6:
                    c.number_format = "$#,##0.00"
                    c.font = Font(name="Arial", size=9, bold=True, color=C_GREEN_DARK)
                    c.alignment = center()
                elif col in (7, 8):
                    c.font = Font(name="Arial", size=9, bold=True,
                                  color=C_GREEN_DARK if save_pct else "888888")
                    c.alignment = center()
                elif col == 9:
                    c.font = Font(name="Arial", size=9, bold=True,
                                  color=C_RED_DARK if sc >= 9 else C_YELLOW_DARK)
                    c.alignment = center()
                elif col == 11:
                    if url:
                        c.hyperlink = url
                    c.font = Font(name="Arial", size=9, bold=True,
                                  color=C_LINK, underline="single")
                    c.alignment = center()
                elif col == 12:
                    c.font = Font(name="Arial", size=9, italic=True, color="444444")
                    c.alignment = left(wrap=True)
                else:
                    c.font = body_font(9)
                    c.alignment = center()
            ws.row_dimensions[ROW].height = 38
            deal_num += 1
            ROW += 1

    ROW += 1

    # ── Section 3: Signal Legend ──────────────────────────────────────────────
    ROW = section_hdr(ws, ROW, "  🔑  SIGNAL KEY")
    legend = [
        ("🔴 HOT BUY",    "Fresh price drop (≤2 days detected) + deep discount (10%+) — act now",        C_RED_BG,     C_RED_DARK),
        ("🟠 Strong",     "Good discount + reasonably fresh — strong sourcing candidate",                  C_YELLOW_BG,  C_YELLOW_DARK),
        ("🟡 Moderate",   "On sale but older deal or shallow discount — keep on radar",                    C_GREY_HDR,   "555555"),
        ("⚪ Watch",      "No discount or deal is stale — monitor only",                                   C_WHITE,      "888888"),
        ("FRESH SCORE",   "0-13 pts: freshness (4) + on sale (2) + discount % (3) + $ saved (2) + rollback/rank (1+1)", C_SUMMARY_ALT, "1F4E79"),
        ("🛒 WM RANK",    "Walmart sales rank for category — lower number = better seller",               C_WHITE,      "000000"),
        ("DEAL AGE",      "🟢 Detected today/recently  🟡 Active 3-7d  🟠 Aging 8-14d  🔴 Old 15d+  — Based on price cache", C_SUMMARY_ALT, "1F4E79"),
        ("⚠️ NOTE",       "Deal Age & Freshness Score improve over time as the price cache builds history. Full accuracy after 2-3 days.", "FFF9E6", "7F6000"),
    ]
    for lbl, desc, bg_c, txt_c in legend:
        ws.merge_cells(f"A{ROW}:C{ROW}")
        ws.merge_cells(f"D{ROW}:{get_column_letter(SUMMARY_NCOLS)}{ROW}")
        c1 = ws.cell(row=ROW, column=1, value=lbl)
        c1.font = Font(name="Arial", size=9, bold=True, color=txt_c)
        c1.fill = fill(bg_c); c1.alignment = left(); c1.border = BORDER_THIN
        c2 = ws.cell(row=ROW, column=4, value=desc)
        c2.font = body_font(9); c2.fill = fill(bg_c)
        c2.alignment = left(); c2.border = BORDER_THIN
        ws.row_dimensions[ROW].height = 18
        ROW += 1


# ── Filter logic ──────────────────────────────────────────────────────────────

def apply_filter(all_data: dict, filter_key: str) -> dict:
    filtered = {}
    for cat_name, cat_data in all_data.items():
        pool     = cat_data.get("pool") or cat_data["products"]
        products = cat_data["products"]

        if filter_key == "full":
            result = products

        elif filter_key == "trending":
            result = cat_data.get("fresh_products") or []
            if not result:
                result = sorted(pool, key=lambda p: signal_score(p), reverse=True)

        elif filter_key == "selling":
            result = [p for p in pool if p.get("best_seller_rank")]
            result.sort(key=lambda p: p.get("best_seller_rank") or 9999)
            if not result:
                result = pool  # fallback if no rank data

        elif filter_key == "on_sale":
            result = [p for p in pool if p.get("onSale")]
            result.sort(key=lambda p: float(p.get("percentSavings") or 0), reverse=True)

        elif filter_key == "hot":
            result = [p for p in pool if signal_score(p) >= 9 and p.get("onSale")]
            result.sort(key=lambda p: signal_score(p), reverse=True)

        else:
            result = products

        filtered[cat_name] = {
            "products":       result,
            "pool":           pool,
            "fresh_products": cat_data.get("fresh_products", []),
        }

    return filtered


# ── Main entry point ──────────────────────────────────────────────────────────

def build_report(all_data: dict, filter_key: str = "full") -> str:
    """Build full Walmart Excel workbook. Returns file path."""
    est = pytz.timezone("US/Eastern")
    now = datetime.now(est)
    ts  = now.strftime("%B %d, %Y  %I:%M %p EST")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    display_data = apply_filter(all_data, filter_key)
    build_summary_sheet(wb, display_data, ts, filter_key=filter_key)

    for cat_name, cat_data in display_data.items():
        build_category_sheet(wb, cat_name, cat_data["products"], ts)

    stamp    = now.strftime("%Y%m%d_%H%M")
    filename = f"WM_Market_Intel_{stamp}.xlsx"
    path     = f"/tmp/{filename}"
    wb.save(path)
    return path
